# Justin Suh
# Python 3.5.1

# importing parts of the openpyxl library
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Fill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# check if at least 4 rows per term./splice point exist so they can be logged
def check_amount(passed_row):
    # counter used to count how many of the same term./splice point exists
    point_count = 0
    # holds the value of current term./splice point
    curr_point = sheet['I' + str(passed_row)].value

    # check iterate through loop until term./splice point changes
    while curr_point == sheet['I' + str(passed_row)].value:
        point_count += 1
        passed_row += 1

    # return value of current row and number of times listed for that term./splice point
    return point_count, passed_row


if load_workbook("zhuntcombo.xlsx"):
    wb = load_workbook("zhuntcombo.xlsx")
    sheet = wb.active

logs = open("Logs.txt", "w")

all_loc_list = []
to_be_del = []

curr_row = 1

while curr_row < sheet.max_row:
    port = sheet['I' + str(curr_row)].value
    cable = sheet['X' + str(curr_row)].value
    port_count, curr_row = check_amount(curr_row)

    for val in range(curr_row - port_count, curr_row):
        oid = sheet['W' + str(val)].value
        combo = str(oid) + "." + str(port) + "." + str(cable) + "." + str(port_count)
        all_loc_list.append(combo)

i = 0
temp_list = []
while i < len(all_loc_list):
    if all_loc_list[i][-2:-1] == ".":
        count = all_loc_list[i][-1:]
    else:
        count = all_loc_list[i][-2:]

    for val in range(i, i + int(count)):
        temp_list.append(all_loc_list[val][:8])

    temp_list.reverse()

    val = 0
    while val < len(all_loc_list):
        #print(val)
        if all_loc_list[val][:8] == temp_list[0]:
            if all_loc_list[val + 1][:8] == temp_list[1]:
                if all_loc_list[val + 3][:8] == temp_list[3]:
                    for j in range(val, val + int(count)):
                        to_be_del.append(all_loc_list[j])
                        del all_loc_list[j]

        if all_loc_list[val][-2:-1] == ".":
            increment = all_loc_list[val][-1:]
        else:
            increment = all_loc_list[val][-2:]

        val += int(increment)

    i += int(count)

val = 0
while val < len(to_be_del):
    logs.write("-------------------------------------\n")
    if all_loc_list[val][-2:-1] == ".":
        count = all_loc_list[val][-1:]
    else:
        count = all_loc_list[val][-2:]

    for i in range(val, val + int(count)):
        logs.write(to_be_del[val] + "\n")

    val += int(count)

logs.close()
