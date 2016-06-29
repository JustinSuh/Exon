# Justin Suh
# Python 3.5.1

# importing parts of the openpyxl library
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Fill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# check if at least 4 rows per term./splice point exist so they can be logged
def check_amount(passed_row):
    # counter used to count how many of the same term./splice point exists
    point_count = 0
    # holds the value of current term./splice point
    curr_point = sheet['B' + str(passed_row)].value

    # check iterate through loop until term./splice point changes
    while curr_point == sheet['B' + str(passed_row)].value:
        point_count += 1
        passed_row += 1

    # return value of current row and number of times listed for that term./splice point
    return point_count, passed_row


# check status of current term./splice point and return
def check_status():
    return sheet['D' + str(curr_row)].value


# find CLLI codes. Location holds index of where clli begins
def locate_clli(passed_string):
    # check for multi-digit ports
    if passed_string[8] == " ":
        location = 12
    else:
        location = 13
    return_clli = passed_string[location:location + 8]
    if return_clli[-1] == " ":
        return_clli = return_clli[:-1]
        return return_clli
    else:
        return return_clli


# find ports. Location holds end of port
def locate_port(passed_string):
    port_begin = 7
    # check for multi-digit ports
    if passed_string[8] == " ":
        location = 8
    else:
        location = 9
    return_port = passed_string[port_begin:location]

    return return_port


# find bay in string for CO. Bay_begin = where beginning of bay in string
def locate_bay(passed_string):
    if len(passed_string) < 11:
        print("too short")
        return_bay = ""

    else:
        if passed_string[8] == " ":
            bay_begin = 21
        else:
            bay_begin = 22
        return_bay = passed_string[bay_begin:bay_begin + 6]

        if return_bay[-1] == " ":
            bay_begin -= 1
            return_bay = passed_string[bay_begin:bay_begin + 6]

    return return_bay


def get_constants():
    count = 0
    for val in range(int(output_curr_row), int(output_row + 1)):
        # constants
        newSheet['A' + str(val)].value = "NATIONAL"
        newSheet['G' + str(val)].value = "SC"
        newSheet['L' + str(val)].value = "LGSS"
        newSheet['M' + str(val)].value = "CCABLE"
        newSheet['O' + str(val)].value = "WORKING"
        newSheet['P' + str(val)].value = ""
        newSheet['V' + str(val)].value = "SC"
        newSheet['I' + str(val)].value = \
            "{} - {}".format(sheet['B' + str(curr_row - final_pair_count)].value,
                             sheet['B' + str(int(curr_row - (final_pair_count / 2)))].value)
        if count == 0:
            newSheet['J' + str(val)].value = \
                sheet['B' + str(curr_row - final_pair_count)].value
            newSheet['K' + str(val)].value = \
                sheet['B' + str(curr_row - final_pair_count)].value
            count = 1
        else:
            newSheet['J' + str(val)].value = \
                sheet['B' + str(int(curr_row - (final_pair_count / 2)))].value
            newSheet['K' + str(val)].value = \
                sheet['B' + str(int(curr_row - (final_pair_count / 2)))].value
            count = 0


# fix non 8 byte clli
def fix_clli(passed_clli):
    clli_begin = correct_co_clli[:-2]
    new_clli = clli_begin + passed_clli[-2:]
    if passed_clli[:3] == correct_co_clli[:3]:
        logs.write("'{}' has been fixed.\n".format(new_clli))
        return new_clli
    else:
        logs.write("Was not able to fix '{}'.\n".format(passed_clli))
        return passed_clli


# place clli's into sheet
def get_clli():
    counter = 0
    increment = 0
    missing_clli = []
    for val in range(int(output_curr_row), int(output_curr_row + 4)):
        newSheet['B' + str(val)].value = correct_co_clli
        if counter < 2:
            clli_info = sheet['H' + str((curr_row - final_pair_count + counter + 1))].value
            clli = locate_clli(clli_info)
            if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append((curr_row - final_pair_count + counter + 1))
            else:
                if len(clli) != 8:
                    logs.write("Line {}: ".format(curr_row - final_pair_count + counter + 1))
                    logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                    clli = fix_clli(clli)
                newSheet['Q' + str(val)].value = clli
            counter += 1
        else:
            clli_info = sheet['H' + str((curr_row - 3 + increment))].value
            clli = locate_clli(clli_info)
            if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append((curr_row - 3 + increment))
            else:
                if len(clli) != 8:
                    logs.write("Line {}: ".format(curr_row - 3 + increment))
                    logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                    clli = fix_clli(clli)
                newSheet['Q' + str(val)].value = clli
            increment += 1

    if final_pair_count > 8:
        iterator = 0
        for val in range(int(output_curr_row + 4), int(output_row + 1)):
            if val % 2 != 1:
                continue
            else:
                # first half of a-pair clli
                clli_info = sheet['H' + str((curr_row - final_pair_count + 1) + iterator)].value
                clli = locate_clli(clli_info)
                if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append(int((curr_row - final_pair_count + 1) + iterator))
                else:
                    if len(clli) != 8:
                        logs.write("Line {}: ".format((curr_row - final_pair_count + 1) + iterator))
                        logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                        clli = fix_clli(clli)
                    newSheet['B' + str(val)].value = clli

                # second half of a-pair clli
                clli_info = sheet['H' + str((curr_row - int(final_pair_count / 2) + 1) + iterator)].value
                clli = locate_clli(clli_info)
                if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append(int((curr_row - int(final_pair_count / 2) + 1) + iterator))
                else:
                    if len(clli) != 8:
                        logs.write("Line {}: ".format((curr_row - int(final_pair_count / 2) + 1) + iterator))
                        logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                        clli = fix_clli(clli)
                    newSheet['B' + str(val + 1)].value = clli

                # first half of z-pair clli
                clli_info = sheet['H' + str((curr_row - final_pair_count + 1) + iterator + 3)].value
                clli = locate_clli(clli_info)
                if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append(int((curr_row - final_pair_count + 1) + iterator + 3))
                else:
                    if len(clli) != 8:
                        logs.write("Line {}: ".format((curr_row - final_pair_count + 1) + iterator + 3))
                        logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                        clli = fix_clli(clli)
                    newSheet['Q' + str(val)].value = clli

                # second half of z-pair clli
                clli_info = sheet['H' + str((curr_row - int(final_pair_count / 2) + 1) + iterator + 3)].value
                clli = locate_clli(clli_info)
                if clli[0:1] != correct_co_clli[0:1]:
                    missing_clli.append(int((curr_row - int(final_pair_count / 2) + 1) + iterator + 3))
                else:
                    if len(clli) != 8:
                        logs.write("Line {}: ".format((curr_row - int(final_pair_count / 2) + 1) + iterator + 3))
                        logs.write("'{}' is not an 8 byte CLLI. Attempting to fix...\n".format(clli))
                        clli = fix_clli(clli)
                    newSheet['Q' + str(val + 1)].value = clli
                iterator += 2
    missing_clli = sorted(set(missing_clli))
    for vals in missing_clli:
        logs.write("Line: {} in the old excel sheet contains a faulty CLLI/is missing a CLLI.\n".format(vals))


# locate ports and place into spreadsheet
def get_port():
    counter = 0
    for val in range(int(output_curr_row), int(output_curr_row + 4)):
        if counter < 2:
            # first half of a-pair port
            port_info = sheet['H' + str((curr_row - final_pair_count))].value
            a_first_port = locate_port(port_info)
            # second half of a-pair port
            port_info = sheet['H' + str((curr_row - int(final_pair_count / 2)))].value
            a_second_port = locate_port(port_info)
            # final port
            a_port = a_first_port + " - " + a_second_port
            newSheet['F' + str(val)].value = a_port

            # first half of z-pair port
            port_info = sheet['H' + str((curr_row - final_pair_count + 1))].value
            z_first_port = locate_port(port_info)
            # second half of z-pair port
            port_info = sheet['H' + str((curr_row - int(final_pair_count / 2) + 1))].value
            z_second_port = locate_port(port_info)
            # final port
            z_port = z_first_port + " - " + z_second_port
            newSheet['U' + str(val)].value = z_port

            counter += 1
        else:
            # first half of a-pair port
            port_info = sheet['H' + str((curr_row - int(final_pair_count / 2) - 1))].value
            a_first_port = locate_port(port_info)
            # second half of a-pair port
            port_info = sheet['H' + str((curr_row - 1))].value
            a_second_port = locate_port(port_info)
            # final port
            a_port = a_first_port + " - " + a_second_port
            newSheet['F' + str(val)].value = a_port

            # first half of z-pair port
            port_info = sheet['H' + str((curr_row - int(final_pair_count / 2) - 2))].value
            z_first_port = locate_port(port_info)
            # second half of z-pair port
            port_info = sheet['H' + str((curr_row - 2))].value
            z_second_port = locate_port(port_info)
            # final port
            z_port = z_first_port + " - " + z_second_port
            newSheet['U' + str(val)].value = z_port

    if final_pair_count > 8:
        iterator = 0
        for val in range(int(output_curr_row + 4), int(output_row + 1)):
            if val % 2 != 1:
                continue
            else:
                # first half of a-pair port
                port_info = sheet['H' + str((curr_row - final_pair_count + 2) + iterator)].value
                a_first_port = locate_port(port_info)
                # second half of a-pair port
                port_info = sheet['H' + str((curr_row - int(final_pair_count / 2) + 2) + iterator)].value
                a_second_port = locate_port(port_info)
                # final port
                a_port = a_first_port + " - " + a_second_port
                newSheet['F' + str(val)].value = a_port
                newSheet['F' + str(val + 1)].value = a_port

                # first half of z-pair port
                port_info = sheet['H' + str((curr_row - final_pair_count + 3) + iterator)].value
                z_first_port = locate_port(port_info)
                # second half of z-pair port
                port_info = sheet['H' + str((curr_row - int(final_pair_count / 2) + 3) + iterator)].value
                z_second_port = locate_port(port_info)
                z_port = z_first_port + " - " + z_second_port
                newSheet['U' + str(val)].value = z_port
                newSheet['U' + str(val + 1)].value = z_port

                iterator += 2


# place bays for CO in new sheet
def get_bay():
    counter = 0
    for val in range(int(output_curr_row), int(output_curr_row + 4)):
        if counter < 2:
            bay_info = sheet['H' + str((curr_row - final_pair_count))].value
            bay = locate_bay(bay_info)
            newSheet['D' + str(val)].value = bay
            counter += 1
        else:
            bay_info = sheet['H' + str((curr_row - int(final_pair_count / 2) - 1))].value
            bay = locate_bay(bay_info)
            newSheet['D' + str(val)].value = bay


# place shelves for locations
def get_shelf():
    x = 1


# place cid
def get_cid():
    counter = 0
    sys_num = ""
    cid = sheet['F' + str((curr_row - final_pair_count))].value
    for char in range(0, len(cid)):
        if cid[char].isdigit() and counter < 3:
            sys_num += cid[char]
            counter += 1
    sys_num = "Sys " + sys_num
    for val in range(int(output_curr_row), int(output_curr_row + 4)):
        if len(sys_num) != 7:
            logs.write("Check Sys number for this pair.")
        else:
            newSheet['P' + str(val)].value = sys_num



# take user input for file name
file_name = input("Enter file name (with extension i.e. 'file.xlsx'): ")

# load excel sheet. Throw error if it can not be opened
print("Opening passed excel file...")
# ************************************BE SURE TO COME BACK TO THIS***********************************
if load_workbook("Madison Ring 1 Revision.xlsx"):
    wb = load_workbook("Madison Ring 1 Revision.xlsx")
# ************************************BE SURE TO COME BACK TO THIS***********************************
    sheet = wb.active
    print("Excel file opened.")
else:
    print("Failed to open file.")
    quit()

print("Opening CBL excel sheet")
if load_workbook("CLEC CBL REPORT 12.07.15.xlsx"):
    alt_wb = load_workbook("CLEC CBL REPORT 12.07.15.xlsx")
    alt_sheet = alt_wb.active
    print("CBL excel file opened.")
else:
    print("Failed to open CBL excel sheet.")
    quit()

print("Creating/opening log file...")
index = file_name.index('.')
# create and open text file for logs
logs = open("Logs - " + file_name[:index] + ".txt", "w")

print("Creating output excel sheet...")
# create output excel sheet
newBook = Workbook()
name = "TIRKS F1 - " + file_name[:index] + ".xlsx"
newSheet = newBook.active
newSheet.title = file_name
newSheet.font = Font(name='Arial', size=8)

# counter for current row in old excel sheet
curr_row = 1
# counter for current and total number of rows in output excel sheet
output_row = 0
output_curr_row = 1
# check if CLLI code is correct
clli_check = False
# check if sys num is correct

# variables to hold co clli code
co_clli = None
incorrect_co_clli = None
correct_co_clli = None

# iterate through excel sheet till you hit the first term./splice point row
while sheet['A' + str(curr_row)].value != "term. point" and sheet['A' + str(curr_row)].value != "splice point":
    curr_row += 1

# iterate through each row of the excel sheet
while curr_row < sheet.max_row:
    first_pair_status = check_status()
    first_pair_count, curr_row = check_amount(curr_row)
    second_pair_status = check_status()
    second_pair_count, curr_row = check_amount(curr_row)
    final_pair_count = first_pair_count + second_pair_count

    logs.write("-------------------------------------------------------------------------------\n")
    # check if at least 8 entries exist
    if final_pair_count <= 7:
        logs.write("{} - {} do not contain enough pairs to be logged.\n"
                   .format(sheet['B' + str(curr_row - final_pair_count)].value,
                           sheet['B' + str(int(curr_row - (final_pair_count / 2)))].value))
    else:
        # check even amount of pairs
        if first_pair_count % 2 != 0:
            logs.write(
                "{} contains an ODD amount of points.\n".format(sheet['B' + str(curr_row - final_pair_count)].value)
            )
        # check even amount of pairs
        if second_pair_count % 2 != 0:
            logs.write(
                "{} contains an ODD amount of points.\n".format(sheet['B' + str(curr_row - final_pair_count / 2)].value)
            )
        # check if ports numbers correlate
        if first_pair_count != second_pair_count:
            logs.write(
                "{} - {} contain an UNEQUAL amount of pairs. Check this pair.\n".format(
                    sheet['B' + str(curr_row - second_pair_count - 1)].value,
                    sheet['B' + str(int(curr_row - 1))].value
                )
            )
            # ************************************BE SURE TO COME BACK TO THIS***********************************

            # ************************************BE SURE TO COME BACK TO THIS***********************************

        else:
            logs.write(
                "{} - {} will be logged.\n".format(
                    sheet['B' + str(curr_row - final_pair_count)].value,
                    sheet['B' + str(int(curr_row - (final_pair_count / 2)))].value
                )
            )
            output_row += final_pair_count / 2

            while not clli_check:
                string = sheet['H' + str(curr_row - final_pair_count)].value
                co_clli = locate_clli(string)
                reply = input("Is '{}' the correct CLLI for the CO at your location? (Y/N): ".format(co_clli))
                if reply == "n" or reply == "N":
                    incorrect_co_clli = co_clli
                    co_clli = input("Please enter the correct CLLI: ")
                    correct_co_clli = co_clli
                else:
                    correct_co_clli = co_clli
                clli_check = True

            if co_clli == incorrect_co_clli:
                co_clli = correct_co_clli

            get_constants()
            get_clli()
            get_port()
            get_bay()
            get_shelf()
            get_cid()

            output_curr_row = output_row + 1

    logs.write("Point {} status: {}\n".format(
        sheet['B' + str(curr_row - second_pair_count - 1)].value, first_pair_status))
    logs.write("Point {} status: {}\n".format(
        sheet['B' + str(int(curr_row - 1))].value, second_pair_status))

print("Logs closed")
logs.close()
print("Output excel closed.")
newBook.save(name)
print("The outputted excel sheet and logs can now be found in the folder where your '{}' is located.\n"
      .format(file_name))

# 7 byte CLLI
# wrong ports
# shelves
# cable names
